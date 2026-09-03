#!/usr/bin/env python3
"""
Fase 6 — Migração do Google Sheets para o Supabase.

Uso:
  python3 scripts/migracao/migrar.py baixar      [--id ID_PLANILHA]   → saida/planilha.xlsx
  python3 scripts/migracao/migrar.py consolidar  [--xlsx ARQ]         → saida/*.json + saida/relatorio.md
  python3 scripts/migracao/migrar.py enviar                            → grava no Supabase (pede confirmação)

Variáveis para `enviar`: SUPABASE_URL, SUPABASE_ANON_KEY, SUPABASE_EMAIL, SUPABASE_SENHA (usuário ADMIN/PCM).

Regras (ver relatório para os números):
  - FILA_OPERACIONAL é a lista-mestre e a fonte confiável da OM (texto). HISTÓRICO_FILA complementa.
  - O status vem da FILA (FINALIZADO/ENCAMINHADO); itens ativos são refinados pelas abas
    PLANEJAMENTO_PCM, ROTEIRO_DIÁRIO, PRÉ_CARGA e RETORNO_PENDÊNCIAS (precedência nessa ordem inversa).
  - OM que virou data/número em outra aba é corrigida pela FILA (mesmo equipamento + patrimônio).
  - CONTROLE_EXECUÇÃO e ESPELHO_PENDÊNCIAS são espelhos sem estado confiável: ignorados (só contados).
  - Duplicatas (equipamento + patrimônio + OM + cliente) são consolidadas.
"""
import argparse, collections, datetime as dt, json, os, re, sys, unicodedata, urllib.request, urllib.parse
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from clientes_mapa import canonico

AQUI = os.path.dirname(os.path.abspath(__file__))
SAIDA = os.path.join(AQUI, 'saida')
ID_PADRAO = '1QyBhlGlIHg4SbQbPI4Jb_zMWzNX1Wu7VbGAzGt3q3rQ'

TIPOS = ['ENTREGA', 'TROCA', 'RETORNO', 'RETORNO AO CLIENTE', 'LOCACAO', 'MANUTENÇÃO', 'RETIRADA', 'DEVOLUÇÃO']
VEICULOS_SEED = {  # placa → nome canônico
    'SRT9D86': 'FIORINO - SRT9D86', 'TTB0J08': 'KIA - TTB0J08', 'TZG3B34': 'SAVEIRO - TZG3B34', 'TTZ7I26': 'KIA - TTZ7I26',
    'SRT9D65': 'FIORINO - SRT9D65', 'TTP8H79': 'SCUDO - TTP8H79', 'SRT9D55': 'STRADA - SRT9D55', 'TTX1H09': 'KIA - TTX1H09',
    'RJM8E69': 'VOLKS - RJM8E69',
}

# ---------------------------------------------------------------- utilitários
def norm(s):
    if s is None: return ''
    s = unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', s).strip().upper()

def texto(v):
    if v is None: return None
    if isinstance(v, float): return str(int(v)) if v.is_integer() else str(v)
    if isinstance(v, int): return str(v)
    if isinstance(v, (dt.datetime, dt.date)): return None   # ver om_de()
    s = str(v).strip().strip("'\"").strip()
    return s or None

def maiusc(v):
    t = texto(v)
    return re.sub(r'\s+', ' ', t).upper() if t else None

def data_iso(v):
    if v is None or v == '': return None
    if isinstance(v, dt.datetime): v = v.date()
    if isinstance(v, dt.date):
        return v.isoformat() if 2000 <= v.year <= 2100 else None
    s = str(v).strip()
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m and 2000 <= int(m[1]) <= 2100: return f'{m[1]}-{m[2]}-{m[3]}'
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if m:
        y = int(m[3]); y = y + 2000 if y < 100 else y
        if 2000 <= y <= 2100: return f'{y:04d}-{int(m[2]):02d}-{int(m[1]):02d}'
    return None

def parece_om_corrompida(v):
    """Célula de OM que o Sheets converteu em data (ou texto de data)."""
    if isinstance(v, (dt.datetime, dt.date)): return True
    s = texto(v) or ''
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m: return 1900 <= int(m[1]) <= 2100          # "1402-01-26" é OM, não data
    return bool(re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}$', s))

def om_recuperada(v):
    """Tenta desfazer a conversão do Sheets sem depender da FILA.
    - datetime/texto com ano < 2000: o número da OM foi lido como serial de data (35112 → 17/02/1996).
    - dd/mm/aaaa com aaaa < 1900: "1446-01" foi lido como jan/1446."""
    d = None
    if isinstance(v, (dt.datetime, dt.date)): d = v if isinstance(v, dt.date) else v.date()
    else:
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', texto(v) or '')
        if m:
            dd, mm, yy = int(m[1]), int(m[2]), int(m[3])
            if yy < 1900: return f'{yy}-{mm:02d}' if dd == 1 else None
            try: d = dt.date(yy, mm, dd)
            except ValueError: return None
    if d is None: return None
    if d.year >= 2000: return None
    serial = (d - dt.date(1899, 12, 30)).days
    return str(serial) if serial > 0 else None

def om_de(v):
    """OM como texto, ou None se corrompida/vazia."""
    if v is None or parece_om_corrompida(v): return None
    t = texto(v)
    return t or None

def tipo_de(v):
    n = norm(v)
    if not n: return 'ENTREGA'
    for t in TIPOS:
        if norm(t) == n: return t
    if n.startswith('LOCA'): return 'LOCACAO'
    if n.startswith('RETORNO AO'): return 'RETORNO AO CLIENTE'
    if n.startswith('RETOR'): return 'RETORNO'
    if n.startswith('RETIR'): return 'RETIRADA'
    if n.startswith('DEVOL'): return 'DEVOLUÇÃO'
    if n.startswith('MANUT'): return 'MANUTENÇÃO'
    if n.startswith('TROC'): return 'TROCA'
    if n.startswith('ENTREG'): return 'ENTREGA'
    return 'ENTREGA'

def veiculo_de(v):
    t = maiusc(v)
    if not t: return None
    placa = re.sub(r'[^A-Z0-9]', '', t)
    for p, nome in VEICULOS_SEED.items():
        if p in placa: return nome
    return t

def quantidade_de(v):
    try: q = float(str(v).replace(',', '.'))
    except (TypeError, ValueError): return 1
    return q if q > 0 else 1

# ---------------------------------------------------------------- leitura
def carregar(xlsx):
    import openpyxl, warnings
    warnings.simplefilter('ignore')
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    abas = {}
    for nome in wb.sheetnames:
        linhas = [r for r in wb[nome].iter_rows(values_only=True) if any(v not in (None, '') for v in r)]
        abas[nome] = linhas
    return abas

def com_cabecalho(linhas):
    """Devolve (índice por nome normalizado, linhas de dados)."""
    if not linhas: return {}, []
    hdr = [norm(h) for h in linhas[0]]
    return {h: i for i, h in enumerate(hdr) if h}, linhas[1:]

def col(row, idx, *nomes):
    for n in nomes:
        i = idx.get(norm(n))
        if i is not None and i < len(row): return row[i]
    return None

# ---------------------------------------------------------------- consolidação
class Demanda(dict):
    pass

def chave_fisica(equip, pat):
    return f'{norm(equip)}|{norm(pat)}'

def consolidar(abas, tecnicos_conhecidos):
    rel = collections.OrderedDict()
    anomalias = collections.defaultdict(list)
    tecn_canon = {norm(t): t for t in tecnicos_conhecidos}
    tecnicos_inativos = set()

    def tecnico_de(v):
        n = norm(v)
        if not n: return None
        if n in tecn_canon: return tecn_canon[n]
        nome = str(v).strip().title()
        anomalias['tecnico_fora_do_config_criado_inativo'].append(nome)
        tecn_canon[n] = nome
        tecnicos_inativos.add(nome)
        return nome

    # --- tipo de controle por equipamento (BASE_TIPO_CONTROLE)
    controle = {}
    idx, dados = com_cabecalho(abas.get('BASE_TIPO_CONTROLE', []))
    for r in dados:
        e = norm(col(r, idx, 'EQUIPAMENTO'))
        if e: controle[e] = (norm(col(r, idx, 'TIPO_CONTROLE')).startswith('PECA'), maiusc(col(r, idx, 'UNIDADE_MEDIDA')))

    demandas = []                      # lista de Demanda
    por_fisica = collections.defaultdict(list)   # equip|pat → demandas
    por_identidade = {}                # equip|pat|om|cliente → demanda

    def nova(**kw):
        d = Demanda(om=None, cliente_nome=None, local=None, tipo='ENTREGA', equipamento_nome=None, patrimonio=None,
                    quantidade=1, unidade=None, tecnico_nome=None, veiculo=None, data_abertura=None, data_planejada=None,
                    data_reagendada=None, status='FILA', status_separacao='NAO_SEPARADO', separado_por=None,
                    data_separacao=None, ordem_parada=None, origem=None, herdado_de_pendencia=False, observacao=None,
                    finalizado_em=None, _fontes=[])
        d.update(kw)
        pat = d['patrimonio']
        if pat and norm(pat) in ('SN', 'S/N', '-', 'N/A'): d['patrimonio'] = None
        e = norm(d['equipamento_nome'])
        if e in controle:
            porq, un = controle[e]
            if porq: d['patrimonio'] = d['patrimonio'] if d['patrimonio'] and not porq else None
            d['unidade'] = d['unidade'] or un
        if d['tipo'] == 'DEVOLUÇÃO EM LOTE': d['tipo'] = 'DEVOLUÇÃO'
        return d

    def identidade(d):
        return f"{norm(d['equipamento_nome'])}|{norm(d['patrimonio'])}|{norm(d['om'])}|{norm(d['cliente_nome'])}"

    def registrar(d, fonte):
        d['_fontes'].append(fonte)
        k = identidade(d)
        if k in por_identidade:
            ex = por_identidade[k]
            if (ex['status'] == 'FINALIZADO' and d['status'] == 'FINALIZADO') or ex['data_abertura'] == d['data_abertura']:
                anomalias['duplicata_consolidada'].append(f"{fonte}: {d['equipamento_nome']} {d['patrimonio'] or ''} OM {d['om']} {d['cliente_nome']}")
                return ex
            anomalias['mesma_identidade_mantida_separada'].append(f"{fonte}: {d['equipamento_nome']} {d['patrimonio'] or ''} OM {d['om']} — {ex['status']} ({ex['data_abertura']}) e {d['status']} ({d['data_abertura']})")
            k = f"{k}#{len(demandas)}"
        por_identidade[k] = d
        demandas.append(d)
        por_fisica[chave_fisica(d['equipamento_nome'], d['patrimonio'])].append(d)
        return d

    # --- 1. FILA_OPERACIONAL (sem cabeçalho; 17 colunas fixas)
    fila = abas.get('FILA_OPERACIONAL', [])
    if fila and norm(fila[0][1]) == 'CLIENTE': fila = fila[1:]
    n_fila_ativas = 0
    for r in fila:
        r = list(r) + [None] * (17 - len(r))
        st = norm(r[11])
        status = 'FINALIZADO' if st == 'FINALIZADO' else ('PRONTO_PARA_PLANEJAR' if st.startswith('PRONTO') else 'ENCAMINHADO' if st == 'ENCAMINHADO' else 'FILA')
        if status != 'FINALIZADO': n_fila_ativas += 1
        d = nova(om=om_de(r[6]), cliente_nome=maiusc(r[1]), local=maiusc(r[2]), tipo=tipo_de(r[3]), equipamento_nome=maiusc(r[4]),
                 patrimonio=texto(r[5]), quantidade=quantidade_de(r[16]), unidade=maiusc(r[15]),
                 data_abertura=data_iso(r[0]), data_planejada=None, status=status, origem=f"MIGRACAO:FILA:{maiusc(r[9]) or ''}",
                 observacao=texto(r[13]), finalizado_em=None)
        if r[6] is not None and d['om'] is None:
            rec = om_recuperada(r[6])
            if rec: d['om'] = rec; anomalias['om_recuperada_por_serial'].append(f"FILA {d['equipamento_nome']} {d['patrimonio']}: {r[6]!r} → {rec}")
            else: anomalias['om_ilegivel_na_fila'].append(f"{d['equipamento_nome']} {d['patrimonio']} → {r[6]!r}")
        if not d['equipamento_nome'] and not d['om']: continue
        registrar(d, 'FILA')
    rel['FILA_OPERACIONAL: linhas'] = len(fila)
    rel['FILA_OPERACIONAL: ativas (não finalizadas)'] = n_fila_ativas

    # --- 2. HISTÓRICO_FILA (finalizadas antigas)
    idx, dados = com_cabecalho(abas.get('HISTÓRICO_FILA', []))
    for r in dados:
        d = nova(om=om_de(col(r, idx, 'OS_CONTRATO')), cliente_nome=maiusc(col(r, idx, 'CLIENTE')), local=maiusc(col(r, idx, 'LOCAL')),
                 tipo=tipo_de(col(r, idx, 'TIPO')), equipamento_nome=maiusc(col(r, idx, 'EQUIPAMENTO')), patrimonio=texto(col(r, idx, 'PATRIMÔNIO')),
                 quantidade=quantidade_de(col(r, idx, 'QUANTIDADE')), unidade=maiusc(col(r, idx, 'UNIDADE_MEDIDA')),
                 data_abertura=data_iso(col(r, idx, 'DATA')), status='FINALIZADO', origem=f"MIGRACAO:HISTORICO_FILA:{maiusc(col(r, idx, 'ORIGEM')) or ''}",
                 observacao=texto(col(r, idx, 'OBS')), finalizado_em=data_iso(col(r, idx, 'DATA_ARQUIVAMENTO')))
        if not d['equipamento_nome'] and not d['om']: continue
        registrar(d, 'HISTORICO_FILA')
    rel['HISTÓRICO_FILA: linhas'] = len(dados)

    # --- OM: corrigir pela FILA quando outra aba trouxe data/número
    def om_corrigida(valor_bruto, equip, pat, cliente):
        om = om_de(valor_bruto)
        cands = por_fisica.get(chave_fisica(equip, pat), [])
        if om is not None:
            return om
        if valor_bruto is None: return None
        # prioridade: candidata ativa, depois mesma cliente, depois qualquer
        ativas = [c for c in cands if c['status'] != 'FINALIZADO' and c['om']]
        mesmo_cli = [c for c in cands if norm(c['cliente_nome']) == norm(cliente) and c['om']]
        for grupo in (ativas, mesmo_cli, [c for c in cands if c['om']]):
            if grupo:
                anomalias['om_corrigida_pela_fila'].append(f"{equip} {pat}: {valor_bruto!r} → {grupo[-1]['om']}")
                return grupo[-1]['om']
        rec = om_recuperada(valor_bruto)
        if rec:
            anomalias['om_recuperada_por_serial'].append(f"{equip} {pat}: {valor_bruto!r} → {rec}")
            return rec
        anomalias['om_corrompida_sem_correcao'].append(f"{equip} {pat} ({cliente}): {valor_bruto!r}")
        return None

    def localizar_ativa(equip, pat, om, cliente, local):
        """Encontra a demanda da FILA correspondente a uma linha de aba ativa."""
        cands = por_fisica.get(chave_fisica(equip, pat), [])
        if pat:
            ativas = [c for c in cands if c['status'] != 'FINALIZADO']
            if len(ativas) == 1: return ativas[0]
            if om:
                por_om = [c for c in ativas if norm(c['om']) == norm(om)] or [c for c in cands if norm(c['om']) == norm(om)]
                if por_om: return por_om[-1]
            if ativas: return ativas[-1]
        else:  # item por quantidade: equipamento + cliente + local
            ativas = [c for c in cands if c['status'] != 'FINALIZADO' and norm(c['cliente_nome']) == norm(cliente) and norm(c['local']) == norm(local)]
            if ativas: return ativas[-1]
            if om:
                por_om = [c for c in cands if norm(c['om']) == norm(om) and norm(c['cliente_nome']) == norm(cliente)]
                if por_om: return por_om[-1]
        return None

    def aplicar_ativa(fonte, equip, pat, om_bruto, cliente, local, tipo, campos):
        pat = texto(pat)
        if pat and norm(pat) in ('SN', 'S/N', '-', 'N/A'): pat = None
        om = om_corrigida(om_bruto, equip, pat, cliente)
        d = localizar_ativa(equip, pat, om, cliente, local)
        if d is None:
            d = registrar(nova(om=om, cliente_nome=maiusc(cliente), local=maiusc(local), tipo=tipo_de(tipo), equipamento_nome=maiusc(equip),
                               patrimonio=texto(pat), origem=f'MIGRACAO:{fonte}', data_abertura=campos.get('data_planejada')), fonte)
            anomalias['ativa_sem_linha_na_fila'].append(f"{fonte}: {equip} {pat or ''} OM {om} {cliente}")
        else:
            d['_fontes'].append(fonte)
            if d['status'] == 'FINALIZADO':
                anomalias['ativa_mas_finalizada_na_fila'].append(f"{fonte}: {equip} {pat or ''} OM {om} — mantida como {campos.get('status')}")
        for k, v in campos.items():
            if v is not None: d[k] = v
        return d

    # --- 3. PLANEJAMENTO_PCM (ativas)
    idx, dados = com_cabecalho(abas.get('PLANEJAMENTO_PCM', []))
    n = 0
    for r in dados:
        st = norm(col(r, idx, 'STATUS PLANO'))
        if st in ('FINALIZADO', 'CANCELADO', ''): continue
        status = 'ROTEIRIZADO' if st == 'ROTEIRIZADO' else 'PLANEJADO' if st == 'PLANEJADO' else 'AGUARDANDO_ROTEIRIZACAO'
        sep = norm(col(r, idx, 'STATUS_EXPEDIÇÃO'))
        aplicar_ativa('PLANEJAMENTO_PCM', col(r, idx, 'EQUIPAMENTO'), col(r, idx, 'PATRIMÔNIO'), col(r, idx, 'OS_CONTRATO'), col(r, idx, 'CLIENTE'),
                      col(r, idx, 'LOCAL'), col(r, idx, 'TIPO'), dict(
            status=status, tecnico_nome=tecnico_de(col(r, idx, 'TÉCNICO')), veiculo=veiculo_de(col(r, idx, 'VEÍCULO')),
            data_planejada=data_iso(col(r, idx, 'DATA EXECUÇÃO')), ordem_parada=int(float(col(r, idx, 'ORDEM'))) if texto(col(r, idx, 'ORDEM')) and re.match(r'^\d+(\.0)?$', texto(col(r, idx, 'ORDEM'))) else None,
            observacao=texto(col(r, idx, 'OBS')), status_separacao='SEPARADO' if sep == 'SEPARADO' else None,
            separado_por=maiusc(col(r, idx, 'SEPARADO_POR')) if sep == 'SEPARADO' else None, data_separacao=data_iso(col(r, idx, 'DATA_SEPARAÇÃO'))))
        n += 1
    rel['PLANEJAMENTO_PCM: ativas aplicadas'] = n

    # --- 4. PRÉ_CARGA (separação do dia)
    linhas = abas.get('PRÉ_CARGA', [])
    idx, dados = com_cabecalho(linhas)
    n = 0
    for r in dados:
        sep = norm(col(r, idx, 'Status Separação'))
        aplicar_ativa('PRE_CARGA', col(r, idx, 'Equipamento'), col(r, idx, 'Patrimônio'), col(r, idx, 'OS'), col(r, idx, 'Cliente'),
                      col(r, idx, 'Local'), col(r, idx, 'Tipo'), dict(
            status='ROTEIRIZADO', tecnico_nome=tecnico_de(col(r, idx, 'Técnico')), veiculo=veiculo_de(col(r, idx, 'Veículo')),
            data_planejada=data_iso(col(r, idx, 'Data')), quantidade=quantidade_de(col(r, idx, 'Qtd')) if col(r, idx, 'Qtd') else None,
            status_separacao='SEPARADO' if sep == 'SEPARADO' else 'NAO_SEPARADO', separado_por=maiusc(col(r, idx, 'Sep. por')), data_separacao=data_iso(col(r, idx, 'Data Sep.'))))
        n += 1
    rel['PRÉ_CARGA: linhas aplicadas'] = n

    # --- 5. ROTEIRO_DIÁRIO (em rota)
    idx, dados = com_cabecalho(abas.get('ROTEIRO_DIÁRIO', []))
    n = 0
    for r in dados:
        st = norm(col(r, idx, 'STATUS_ROTA'))
        status = 'EM_DESLOCAMENTO' if st.startswith('EM DESLOC') else 'AGUARDANDO_SAIDA' if st.startswith('AGUARDANDO') else 'ROTEIRIZADO'
        sep = norm(col(r, idx, 'STATUS_EXPEDIÇÃO'))
        ordem = texto(col(r, idx, 'ORDEM'))
        aplicar_ativa('ROTEIRO_DIARIO', col(r, idx, 'EQUIPAMENTO'), col(r, idx, 'PATRIMÔNIO'), col(r, idx, 'Coluna 10', 'OM/CONTRATO'), col(r, idx, 'CLIENTE'),
                      col(r, idx, 'LOCAL'), col(r, idx, 'TIPO'), dict(
            status=status, tecnico_nome=tecnico_de(col(r, idx, 'TÉCNICO')), veiculo=veiculo_de(col(r, idx, 'VEÍCULO')),
            data_planejada=data_iso(col(r, idx, 'DATA')), ordem_parada=int(float(ordem)) if ordem and re.match(r'^\d+(\.0)?$', ordem) else None,
            status_separacao='SEPARADO' if sep in ('SEPARADO', 'LIBERADO PARA ROTA') else 'NAO_SEPARADO',
            separado_por=maiusc(col(r, idx, 'SEPARADO_POR')), data_separacao=data_iso(col(r, idx, 'DATA_SEPARAÇÃO'))))
        n += 1
    rel['ROTEIRO_DIÁRIO: linhas aplicadas'] = n

    # --- 6. RETORNO_PENDÊNCIAS (pendências aguardando retorno → planejamento com data reagendada)
    idx, dados = com_cabecalho(abas.get('RETORNO_PENDÊNCIAS', []))
    n = 0
    for r in dados:
        if norm(col(r, idx, 'STATUS_RETORNO')) == 'RESOLVIDO': continue
        reag = data_iso(col(r, idx, 'REAGENDAR_PARA'))
        obs = ' · '.join(x for x in [texto(col(r, idx, 'OBS')), f"pendência de {data_iso(col(r, idx, 'DATA_EXECUÇÃO'))}", f"retorno: {texto(col(r, idx, 'REAGENDAR_PARA'))}" if not reag and texto(col(r, idx, 'REAGENDAR_PARA')) else None] if x)
        aplicar_ativa('RETORNO_PENDENCIAS', col(r, idx, 'EQUIPAMENTO'), col(r, idx, 'PATRIMÔNIO'), col(r, idx, 'OM/CONTRATO'), col(r, idx, 'CLIENTE'),
                      col(r, idx, 'LOCAL'), col(r, idx, 'TIPO'), dict(
            status='AGUARDANDO_ROTEIRIZACAO', tecnico_nome=tecnico_de(col(r, idx, 'TÉCNICO')), veiculo=veiculo_de(col(r, idx, 'VEÍCULO')),
            data_planejada=reag, data_reagendada=reag, herdado_de_pendencia=True, observacao=obs or None, ordem_parada=None))
        n += 1
    rel['RETORNO_PENDÊNCIAS: aplicadas'] = n

    # --- 7. Enriquecer finalizadas com HISTÓRICO_PCM (técnico/veículo/data) — colunas às vezes deslocadas
    STATUS_FINAIS = {'FINALIZADO', 'CANCELADO', 'PENDENTE'}
    linhas = abas.get('HISTÓRICO_PCM', [])
    enriq = 0
    for r in linhas[1:]:
        r = list(r) + [None] * (17 - len(r))
        if norm(r[9]) in STATUS_FINAIS or (norm(r[6]) not in {norm(t) for t in TIPOS} and norm(r[10]) not in STATUS_FINAIS):
            # layout antigo: DATA_CONCLUSÃO, DATA_EXECUÇÃO, TÉCNICO, VEÍCULO, LOCAL, CLIENTE, EQUIPAMENTO, PAT, OM, STATUS_FINAL, RESULTADO
            concl, exec_, tec, vei, cli, equip, pat, om_b, resultado = r[0], r[1], r[2], r[3], r[5], r[6], r[7], r[8], r[10]
        else:
            # layout novo: ..., CLIENTE, TIPO, EQUIPAMENTO, PAT, OM, STATUS_FINAL, RESULTADO
            concl, exec_, tec, vei, cli, equip, pat, om_b, resultado = r[0], r[1], r[2], r[3], r[5], r[7], r[8], r[9], r[11]
        cands = por_fisica.get(chave_fisica(equip, pat), [])
        if not cands: continue
        om = om_de(om_b)
        alvo = None
        if om:
            for c in cands:
                if norm(c['om']) == norm(om): alvo = c
        if alvo is None and len(cands) == 1 and pat: alvo = cands[0]
        if alvo is None or alvo['status'] != 'FINALIZADO': continue
        if not alvo['tecnico_nome']: alvo['tecnico_nome'] = tecnico_de(tec)
        if not alvo['veiculo']: alvo['veiculo'] = veiculo_de(vei)
        if not alvo['data_planejada']: alvo['data_planejada'] = data_iso(exec_)
        if not alvo['finalizado_em']: alvo['finalizado_em'] = data_iso(concl)
        if resultado and norm(resultado) not in ('FINALIZADO', 'CONCLUIDO', 'EXECUTADO', ''):
            alvo['observacao'] = ' · '.join(x for x in [alvo['observacao'], f'resultado: {texto(resultado)}'] if x)
        enriq += 1
    rel['HISTÓRICO_PCM: finalizadas enriquecidas'] = enriq

    # --- espelhos ignorados (só contagem)
    rel['CONTROLE_EXECUÇÃO: ignorada (espelho sem resultado), linhas'] = max(0, len(abas.get('CONTROLE_EXECUÇÃO', [])) - 1)
    rel['ESPELHO_PENDÊNCIAS: ignorada (espelho), linhas'] = max(0, len(abas.get('ESPELHO_PENDÊNCIAS', [])) - 1)
    rel['HISTÓRICO_ROTEIROS/EXPEDIÇÃO/PLANEJAMENTO: não importadas (redundantes com FILA + HISTÓRICO_PCM)'] = ''

    # --- finalizado_em padrão
    for d in demandas:
        if d['status'] == 'FINALIZADO' and not d['finalizado_em']:
            d['finalizado_em'] = d['data_planejada'] or d['data_abertura']
        if d['status'] in ('ROTEIRIZADO', 'AGUARDANDO_SAIDA', 'EM_DESLOCAMENTO') and (not d['tecnico_nome'] or not d['data_planejada']):
            anomalias['em_rota_sem_tecnico_ou_data'].append(f"{d['equipamento_nome']} {d['patrimonio'] or ''} OM {d['om']} ({d['status']})")

    # --- clientes: nome canônico + apelidos (grafias originais)
    apelidos = collections.defaultdict(set)
    for d in demandas:
        if not d['cliente_nome']: continue
        canon = canonico(d['cliente_nome'])
        if norm(canon) != norm(d['cliente_nome']): apelidos[canon].add(d['cliente_nome'])
        else: apelidos.setdefault(canon, set())
        d['cliente_nome'] = canon
    rel['Clientes: grafias distintas na planilha'] = sum(len(v) for v in apelidos.values()) + len(apelidos)
    clientes = [dict(nome=k, apelidos=sorted(v)) for k, v in sorted(apelidos.items())]

    # --- cadastros derivados
    equip = {}
    for d in demandas:
        if not d['equipamento_nome']: continue
        k = (d['equipamento_nome'], d['patrimonio'])
        porq = controle.get(norm(d['equipamento_nome']), (False, None))[0]
        equip[k] = dict(nome=d['equipamento_nome'], patrimonio=None if porq else d['patrimonio'], controlado_por_quantidade=porq, unidade=d['unidade'] or ('UNIDADE' if not porq else 'PEÇA'))
    equipamentos = list({(e['nome'], e['patrimonio']): e for e in equip.values()}.values())
    veiculos = sorted({d['veiculo'] for d in demandas if d['veiculo']})
    tecnicos = sorted({d['tecnico_nome'] for d in demandas if d['tecnico_nome']})

    # sugestões de apelidos: clientes que compartilham o início do nome
    grupos = collections.defaultdict(list)
    for c in [c['nome'] for c in clientes]:
        raiz = ' '.join(w for w in norm(c).replace('CONSTRUTORA', '').replace('CONSORCIO', '').split()[:1])
        if raiz: grupos[raiz].append(c)
    sugest = {k: v for k, v in grupos.items() if len(v) > 1}
    for k, v in sorted(sugest.items(), key=lambda x: -len(x[1]))[:25]:
        anomalias['clientes_possivelmente_iguais'].append(f"{k}: " + ' | '.join(v[:8]) + (f' | … +{len(v)-8}' if len(v) > 8 else ''))

    return demandas, dict(clientes=clientes, equipamentos=equipamentos, veiculos=veiculos, tecnicos=tecnicos, tecnicos_inativos=sorted(tecnicos_inativos)), rel, anomalias

# ---------------------------------------------------------------- relatório
def escrever_relatorio(demandas, cad, rel, anomalias, caminho):
    st = collections.Counter(d['status'] for d in demandas)
    tec = collections.Counter(d['tecnico_nome'] or '— sem técnico' for d in demandas if d['status'] not in ('FINALIZADO', 'CANCELADO'))
    linhas = ['# Relatório da consolidação', '', f'Gerado em {dt.datetime.now():%d/%m/%Y %H:%M}', '', '## Leitura das abas', '']
    for k, v in rel.items(): linhas.append(f'- {k}: {v}')
    linhas += ['', '## Resultado', '', f'- **Demandas consolidadas: {len(demandas)}**', f"- Ativas: {sum(v for k, v in st.items() if k not in ('FINALIZADO', 'CANCELADO'))} · Finalizadas: {st['FINALIZADO']}", '', '### Por status', '']
    for k, v in sorted(st.items(), key=lambda x: -x[1]): linhas.append(f'- {k}: {v}')
    linhas += ['', '### Ativas por técnico', '']
    for k, v in tec.most_common(): linhas.append(f'- {k}: {v}')
    linhas += ['', '### Cadastros derivados', '', f"- Clientes distintos: {len(cad['clientes'])}", f"- Equipamentos (nome + patrimônio): {len(cad['equipamentos'])}", f"- Veículos: {', '.join(cad['veiculos'])}", f"- Técnicos: {', '.join(cad['tecnicos'])}", '']
    linhas += ['## Anomalias (o que a planilha tinha de errado e o que foi feito)', '']
    linhas += ['### Clientes unificados (canônico ← grafias)', '']
    for c in sorted(cad['clientes'], key=lambda c: -len(c['apelidos']))[:40]:
        if c['apelidos']: linhas.append(f"- **{c['nome']}** ← {' | '.join(c['apelidos'])}")
    linhas.append('')
    for k, lst in sorted(anomalias.items(), key=lambda x: -len(x[1])):
        linhas.append(f'### {k}: {len(lst)}'); linhas.append('')
        for x in lst[:15]: linhas.append(f'- {x}')
        if len(lst) > 15: linhas.append(f'- … e mais {len(lst) - 15}')
        linhas.append('')
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    open(caminho, 'w', encoding='utf-8').write('\n'.join(linhas))

# ---------------------------------------------------------------- envio
def rest(url, key, token, metodo, caminho, corpo=None, prefer=None):
    req = urllib.request.Request(f'{url}/rest/v1/{caminho}', method=metodo, data=json.dumps(corpo).encode() if corpo is not None else None,
                                 headers={'apikey': key, 'Authorization': f'Bearer {token}', 'Content-Type': 'application/json', **({'Prefer': prefer} if prefer else {})})
    with urllib.request.urlopen(req, timeout=120) as r:
        t = r.read().decode()
        return json.loads(t) if t else None

def enviar(saida):
    url, key, email, senha = (os.environ.get(k) for k in ('SUPABASE_URL', 'SUPABASE_ANON_KEY', 'SUPABASE_EMAIL', 'SUPABASE_SENHA'))
    if not all([url, key, email, senha]): sys.exit('Defina SUPABASE_URL, SUPABASE_ANON_KEY, SUPABASE_EMAIL e SUPABASE_SENHA.')
    url = url.rstrip('/')
    req = urllib.request.Request(f'{url}/auth/v1/token?grant_type=password', data=json.dumps({'email': email, 'password': senha}).encode(), headers={'apikey': key, 'Content-Type': 'application/json'})
    token = json.load(urllib.request.urlopen(req))['access_token']
    demandas = json.load(open(os.path.join(saida, 'demandas.json'), encoding='utf-8'))
    cad = json.load(open(os.path.join(saida, 'cadastros.json'), encoding='utf-8'))

    existentes = rest(url, key, token, 'GET', 'demandas?select=id&limit=1')
    print(f'Banco: {"JÁ TEM demandas" if existentes else "sem demandas"}. Vou inserir {len(demandas)} demandas, '
          f'{len(cad["clientes"])} clientes, {len(cad["equipamentos"])} equipamentos.')
    if input('Confirma? (digite SIM) ') != 'SIM': sys.exit('Cancelado.')

    # técnicos e veículos (cria os que faltam)
    tecs = {norm(t['nome']): t['id'] for t in rest(url, key, token, 'GET', 'tecnicos?select=id,nome')}
    for nome in cad['tecnicos']:
        if norm(nome) not in tecs:
            t = rest(url, key, token, 'POST', 'tecnicos', [{'nome': nome, 'cor': '#64748b', 'ativo': nome not in cad.get('tecnicos_inativos', [])}], 'return=representation')[0]
            tecs[norm(nome)] = t['id']; print('técnico criado:', nome, '(inativo)' if nome in cad.get('tecnicos_inativos', []) else '')
    veics = {norm(v['nome']) for v in rest(url, key, token, 'GET', 'veiculos?select=nome')}
    novos = [{'nome': v, 'placa': (re.search(r'[A-Z]{3}\d[A-Z0-9]\d{2}', v.replace(' ', '')) or [None])[0]} for v in cad['veiculos'] if norm(v) not in veics]
    if novos: rest(url, key, token, 'POST', 'veiculos', novos, 'return=minimal'); print('veículos criados:', len(novos))
    # clientes: resolve por nome/apelido, cria os que faltam
    cli_rows = rest(url, key, token, 'GET', 'clientes?select=id,nome,apelidos')
    cli = {}
    for c in cli_rows:
        cli[norm(c['nome'])] = c['id']
        for a in c.get('apelidos') or []: cli[norm(a)] = c['id']
    faltam, atualizar = [], []
    for c in cad['clientes']:
        cid = cli.get(norm(c['nome']))
        if not cid: faltam.append({'nome': c['nome'], 'apelidos': c['apelidos']}); continue
        ex = next(x for x in cli_rows if x['id'] == cid)
        novos = [a for a in c['apelidos'] if norm(a) not in {norm(x) for x in (ex.get('apelidos') or [])} and norm(a) != norm(ex['nome'])]
        if novos: atualizar.append((cid, sorted(set((ex.get('apelidos') or []) + novos))))
        for a in c['apelidos']: cli[norm(a)] = cid
        cli[norm(c['nome'])] = cid
    for i in range(0, len(faltam), 200):
        for c in rest(url, key, token, 'POST', 'clientes', faltam[i:i + 200], 'return=representation'):
            cli[norm(c['nome'])] = c['id']
            for a in c.get('apelidos') or []: cli[norm(a)] = c['id']
    for cid, aps in atualizar: rest(url, key, token, 'PATCH', f'clientes?id=eq.{cid}', {'apelidos': aps}, 'return=minimal')
    print('clientes criados:', len(faltam), '· apelidos completados em', len(atualizar))
    # equipamentos
    eq_rows = rest(url, key, token, 'GET', 'equipamentos?select=id,nome,patrimonio&limit=10000')
    eq = {(norm(e['nome']), norm(e['patrimonio'])): e['id'] for e in eq_rows}
    faltam = [e for e in cad['equipamentos'] if (norm(e['nome']), norm(e['patrimonio'])) not in eq]
    for i in range(0, len(faltam), 200):
        for e in rest(url, key, token, 'POST', 'equipamentos', faltam[i:i + 200], 'return=representation'): eq[(norm(e['nome']), norm(e['patrimonio']))] = e['id']
    print('equipamentos criados:', len(faltam))
    # demandas
    lote = []
    for d in demandas:
        row = {k: v for k, v in d.items() if not k.startswith('_') and k != 'tecnico_nome'}
        row['tecnico_id'] = tecs.get(norm(d['tecnico_nome'])) if d['tecnico_nome'] else None
        row['cliente_id'] = cli.get(norm(d['cliente_nome']))
        row['equipamento_id'] = eq.get((norm(d['equipamento_nome']), norm(d['patrimonio'])))
        if row['finalizado_em'] and len(row['finalizado_em']) == 10: row['finalizado_em'] += 'T12:00:00Z'
        lote.append(row)
    for i in range(0, len(lote), 300):
        rest(url, key, token, 'POST', 'demandas', lote[i:i + 300], 'return=minimal')
        print(f'demandas: {min(i + 300, len(lote))}/{len(lote)}')
    print('Concluído.')

# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('modo', choices=['baixar', 'consolidar', 'enviar'])
    ap.add_argument('--id', default=ID_PADRAO); ap.add_argument('--xlsx', default=os.path.join(SAIDA, 'planilha.xlsx'))
    a = ap.parse_args()
    os.makedirs(SAIDA, exist_ok=True)
    if a.modo == 'baixar':
        urllib.request.urlretrieve(f'https://docs.google.com/spreadsheets/d/{a.id}/export?format=xlsx', a.xlsx)
        print('baixado:', a.xlsx, os.path.getsize(a.xlsx), 'bytes')
    elif a.modo == 'consolidar':
        abas = carregar(a.xlsx)
        tecs = [texto(r[0]) for r in abas.get('CONFIG_TECNICOS', [])[1:] if texto(r[0])]
        demandas, cad, rel, anomalias = consolidar(abas, tecs)
        json.dump([{k: v for k, v in d.items() if k != '_fontes'} for d in demandas], open(os.path.join(SAIDA, 'demandas.json'), 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
        json.dump(cad, open(os.path.join(SAIDA, 'cadastros.json'), 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
        escrever_relatorio(demandas, cad, rel, anomalias, os.path.join(SAIDA, 'relatorio.md'))
        print(open(os.path.join(SAIDA, 'relatorio.md'), encoding='utf-8').read())
    else:
        enviar(SAIDA)

if __name__ == '__main__':
    main()
